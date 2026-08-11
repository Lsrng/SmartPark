# -*- coding: utf-8 -*-
"""
物业账单 Excel 测试数据生成器
生成符合 BillExcelRowDTO 结构的 .xlsx 文件，用于 JMeter 压测

依赖安装: pip install openpyxl
运行: python generate_excel.py [output_dir]
"""

import openpyxl
import os
import sys
from datetime import date, timedelta
import random


# Excel 列头（与 BillExcelRowDTO 的 @ExcelProperty 注解对应）
HEADERS = [
    "楼栋号", "单元号", "房号", "业主姓名", "联系电话",
    "费用类型", "计费起始日期", "计费截止日期", "应收金额",
    "缴费截止日期", "备注"
]

# 数据池
BUILDINGS = [f"A栋", f"B栋", f"C栋", f"D栋", f"E栋", f"F栋", f"G栋", f"H栋"]
UNITS = [f"1单元", f"2单元", f"3单元", f"4单元"]
SURNAMES = list("赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜")
GIVEN_NAMES = ["伟", "芳", "娜", "敏", "静", "强", "磊", "军", "洋", "勇",
               "艳", "杰", "娟", "涛", "明", "超", "秀英", "霞", "平", "刚",
               "桂英", "建华", "志强", "丽娟", "晓明", "雅婷", "昊天", "思远"]
FEE_TYPES = ["物业费", "水费", "电费", "停车费", "垃圾清运费", "维修费", "其他"]
REMARKS = ["正常缴费", "逾期补缴", "分期缴费", "减免申请", "代收代缴", ""]


def random_phone():
    """生成合法手机号（1开头，11位）"""
    prefixes = ["138", "139", "136", "137", "150", "151", "158", "159", "188", "189", "156", "186"]
    prefix = random.choice(prefixes)
    suffix = "".join([str(random.randint(0, 9)) for _ in range(8)])
    return prefix + suffix


def random_name():
    """生成随机中文姓名"""
    surname = random.choice(SURNAMES)
    given = random.choice(GIVEN_NAMES)
    return surname + given


def generate_row(row_index):
    """生成一行合法的账单数据"""
    building = random.choice(BUILDINGS)
    unit = random.choice(UNITS)
    room = f"{random.randint(1, 30):02d}0{random.randint(1, 9)}"
    owner = random_name()
    phone = random_phone()
    fee_type = random.choice(FEE_TYPES)

    # 计费日期：起始 2025-01-01 到 2025-12-01 之间随机
    start_date = date(2025, random.randint(1, 12), random.randint(1, 28))
    end_date = start_date + timedelta(days=random.randint(30, 365))
    # 缴费截止日期 >= 计费截止日期
    due_date = end_date + timedelta(days=random.randint(1, 90))

    # 金额：100~5000，最多2位小数
    amount = round(random.uniform(100, 5000), 2)
    remark = random.choice(REMARKS)

    # 日期列使用 date 对象（非字符串），EasyExcel 直接解析为 LocalDate
    # 金额列使用数字（非字符串），EasyExcel 直接解析为 BigDecimal
    return [
        building, unit, room, owner, phone,
        fee_type,
        start_date,
        end_date,
        amount,
        due_date,
        remark
    ]


def generate_excel(file_path, row_count):
    """生成指定行数的 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "账单数据"

    # 写入表头
    ws.append(HEADERS)

    # 批量写入数据
    batch_size = 5000
    for batch_start in range(0, row_count, batch_size):
        batch_end = min(batch_start + batch_size, row_count)
        rows = []
        for i in range(batch_start, batch_end):
            rows.append(generate_row(i))
        for row in rows:
            ws.append(row)

    # 设置列宽
    column_widths = [12, 10, 10, 12, 15, 12, 15, 15, 12, 15, 20]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(file_path)
    size_kb = os.path.getsize(file_path) / 1024
    print(f"生成: {file_path} | 行数: {row_count:,} | 大小: {size_kb:.0f} KB")


def main():
    # 输出目录
    if len(sys.argv) > 1:
        output_dir = sys.argv[1]
    else:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test-files")

    os.makedirs(output_dir, exist_ok=True)
    print(f"输出目录: {output_dir}")
    print(f"开始生成测试数据...\n")

    # 生成不同规模的 Excel 文件
    configs = [
        ("bill-1k-50kb.xlsx", 1000),       # 小规模：基线测试
        ("bill-5k-250kb.xlsx", 5000),      # 小中规模
        ("bill-10k-500kb.xlsx", 10000),    # 标准场景
        ("bill-20k-1mb.xlsx", 20000),      # 中等压力
        ("bill-50k-2.5mb.xlsx", 50000),    # 高压测试
        ("bill-100k-5mb.xlsx", 100000),    # 边界上限（~5MB）
    ]

    for filename, row_count in configs:
        file_path = os.path.join(output_dir, filename)
        generate_excel(file_path, row_count)

    # 额外生成边界/异常测试文件
    print(f"\n生成边界测试文件...")

    # 6MB 超限文件（约 12 万行）
    oversize_path = os.path.join(output_dir, "bill-oversize-6mb.xlsx")
    generate_excel(oversize_path, 120000)

    # 小文件（100行）
    tiny_path = os.path.join(output_dir, "bill-tiny-100rows.xlsx")
    generate_excel(tiny_path, 100)

    # 空数据文件（只有表头）
    empty_path = os.path.join(output_dir, "bill-empty.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "账单数据"
    ws.append(HEADERS)
    wb.save(empty_path)
    print(f"生成: {empty_path} | 行数: 0 (仅表头)")

    print(f"\n{'='*50}")
    print(f"全部文件生成完毕！")
    print(f"输出目录: {output_dir}")
    print(f"{'='*50}")


if __name__ == "__main__":
    # 设置随机种子保证可复现
    random.seed(42)
    main()